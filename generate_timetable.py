import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

src = r'E:\AppRoutine\timetable_reports_2_20260712_2108_student.xlsx'
out = r'E:\AppRoutine\CS_Sections_Timetable_Upto_2PM.xlsx'

wb = openpyxl.load_workbook(src, data_only=True)
ws = wb['Section Grid']

# P1 through P6 are columns 3,4,5,6,7,8
PERIOD_COLS = list(range(3, 9))  # P1(3) through P6(8)
HEADERS = ['Section', 'Day', 'P1 (8-9 AM)', 'P2 (9-10 AM)', 'P3 (10-11 AM)', 'P4 (11-12 PM)', 'P5 (12-1 PM)', 'P6 (1-2 PM)']

sections_data = []
i = 2
while i <= ws.max_row:
    cell = ws.cell(row=i, column=1).value
    cell2 = ws.cell(row=i, column=2).value
    if cell and 'Sem 5' in str(cell) and 'CS' in str(cell) and cell2 and 'course' in str(cell2):
        sec_name = str(cell).split('|')[-1].strip()
        days = []
        for d in range(1, 6):
            r = i + d
            day_name = ws.cell(row=r, column=2).value
            periods = []
            for p_col in PERIOD_COLS:
                val = ws.cell(row=r, column=p_col).value or ''
                periods.append(val)
            days.append({'day': day_name, 'periods': periods})
        sections_data.append({'section': sec_name, 'days': days})
        i += 6
    else:
        i += 1

owb = openpyxl.Workbook()
ows = owb.active
ows.title = 'CS Timetable Upto 2PM'

header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
day_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
sec_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
left_align = Alignment(wrap_text=True, vertical='center', horizontal='left')

col_widths = [12, 12, 30, 30, 30, 30, 30, 30]

for ci, h in enumerate(HEADERS, 1):
    c = ows.cell(row=1, column=ci, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center_align
    c.border = thin_border
    ows.column_dimensions[get_column_letter(ci)].width = col_widths[ci-1]

row_num = 2
for sd in sections_data:
    # Section header row (merged across all columns)
    c = ows.cell(row=row_num, column=1, value=sd['section'])
    c.font = Font(bold=True, size=11, color='2F5496')
    c.fill = sec_fill
    c.alignment = center_align
    c.border = thin_border
    for ci in range(2, len(HEADERS)+1):
        c = ows.cell(row=row_num, column=ci)
        c.fill = sec_fill
        c.border = thin_border
    ows.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=len(HEADERS))
    c = ows.cell(row=row_num, column=2, value='Sem 5 CS - Timetable (8 AM to 2 PM)')
    c.font = Font(bold=True, size=10, italic=True, color='2F5496')
    c.alignment = center_align
    c.fill = sec_fill
    c.border = thin_border
    row_num += 1

    for dd in sd['days']:
        c = ows.cell(row=row_num, column=1, value=sd['section'])
        c.font = Font(size=10)
        c.alignment = center_align
        c.border = thin_border

        c = ows.cell(row=row_num, column=2, value=dd['day'])
        c.font = Font(bold=True, size=10)
        c.fill = day_fill
        c.alignment = center_align
        c.border = thin_border

        for pi, val in enumerate(dd['periods'], 3):
            c = ows.cell(row=row_num, column=pi, value=val)
            c.alignment = left_align
            c.border = thin_border
            c.font = Font(size=9)
        row_num += 1

ows.auto_filter.ref = f"A1:H{row_num-1}"
ows.freeze_panes = 'A2'

owb.save(out)
print(f'Done! Timetable saved to: {out}')
print(f'Total CS sections: {len(sections_data)}')
